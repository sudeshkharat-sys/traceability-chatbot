import io
import json
import logging
import re
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

from app.connectors.state_db_connector import StateDBConnector
from app.connectors.database import get_connector
from app.queries import InputRecordQueries
import backend.models.schemas.z_stage_schemas as schemas

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/input", tags=["input"])

# ── Strict allowed values for constrained columns ─────────────────────────────
ALLOWED_TYPE        = {"WH", "USV"}
ALLOWED_RYG         = {"R", "Y", "G"}
ALLOWED_ATTRI       = {"M&M Design", "M&M process", "Supplier Design", "Supplier Process", "Under Analysis"}
ALLOWED_Z_E         = {"Z", "E"}
ALLOWED_ATTRIBUTION = {"M", "P", "D", "U"}
ALLOWED_STATUS_3M   = {"R", "G"}

# Case-insensitive lookup maps — normalise incoming values to the canonical form
_ATTRI_NORM = {v.lower(): v for v in ALLOWED_ATTRI}


def _normalise_attri(val: str | None) -> str | None:
    """Return the canonical Attri. value regardless of how the user typed it."""
    if val is None:
        return None
    return _ATTRI_NORM.get(val.strip().lower(), val.strip())


def _validate_row(rec: dict) -> str | None:
    """Return a human-readable reason string if the row is invalid, else None."""
    issues = []
    if rec.get("type") is not None and rec["type"] not in ALLOWED_TYPE:
        issues.append(f"Type '{rec['type']}' not in {sorted(ALLOWED_TYPE)}")
    if rec.get("ryg") is not None and rec["ryg"] not in ALLOWED_RYG:
        issues.append(f"RYG '{rec['ryg']}' not in {sorted(ALLOWED_RYG)}")
    if rec.get("attri") is not None and rec["attri"] not in ALLOWED_ATTRI:
        issues.append(f"Attri. '{rec['attri']}' not in allowed values")
    if rec.get("z_e") is not None and rec["z_e"] not in ALLOWED_Z_E:
        issues.append(f"Z/E '{rec['z_e']}' not in {sorted(ALLOWED_Z_E)}")
    if rec.get("attribution") is not None and rec["attribution"] not in ALLOWED_ATTRIBUTION:
        issues.append(f"Attribution '{rec['attribution']}' not in {sorted(ALLOWED_ATTRIBUTION)}")
    if rec.get("status_3m") is not None and rec["status_3m"] not in ALLOWED_STATUS_3M:
        issues.append(f"Status(3M) '{rec['status_3m']}' not in {sorted(ALLOWED_STATUS_3M)}")
    return "; ".join(issues) if issues else None


# Month keys expected in the Excel (Jan 2024 → Mar 2026)
# Used only for the download endpoint header; upload detects months dynamically.
MONTHLY_KEYS = [
    "2024-01", "2024-02", "2024-03", "2024-04", "2024-05", "2024-06",
    "2024-07", "2024-08", "2024-09", "2024-10", "2024-11", "2024-12",
    "2025-01", "2025-02", "2025-03", "2025-04", "2025-05", "2025-06",
    "2025-07", "2025-08", "2025-09", "2025-10", "2025-11", "2025-12",
    "2026-01", "2026-02", "2026-03",
]


def _row_to_dict(row) -> dict:
    return dict(row._mapping)


def _clean_text(s: str) -> str:
    """Remove characters that cannot be stored in WIN1252 (e.g. zero-width spaces)."""
    s = re.sub(r'[\u200b\u200c\u200d\u200e\u200f\ufeff\u00ad]', '', s)
    s = s.encode('cp1252', errors='replace').decode('cp1252')
    return s


def _safe_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if s.startswith("="):
        return None
    s = _clean_text(s)
    return s or None


def _safe_int(value) -> int | None:
    if value is None:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _safe_date(value) -> str | None:
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if s.startswith("=") or not s or s == "00:00:00":
        return None
    return s


def _parse_excel(file_bytes: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes), data_only=True, read_only=True
    )
    ws = wb.active

    # ── Detect column positions from header row ───────────────────────────────
    # Month columns are datetime objects in the header.
    # Trailing columns are matched by header text.
    monthly_col_map: dict[int, str] = {}   # col_index → "YYYY-MM"
    field_defect_col: int = 45             # fallback default
    status_3m_col: int = 46               # fallback default

    rows_iter = ws.iter_rows(values_only=True)
    first_row = next(rows_iter, None)
    if first_row:
        for col_idx, cell_val in enumerate(first_row):
            if hasattr(cell_val, "year"):  # datetime → monthly column
                key = f"{cell_val.year}-{str(cell_val.month).zfill(2)}"
                monthly_col_map[col_idx] = key
            elif isinstance(cell_val, str):
                clean = cell_val.strip().lower().replace("\n", " ")
                if "field defect" in clean:
                    field_defect_col = col_idx
                elif "status" in clean and "3" in clean:
                    status_3m_col = col_idx

    records = []
    for row in rows_iter:
        if not any(v is not None for v in row):
            continue  # skip empty rows

        # Monthly data: use dynamic column map from header
        monthly = {}
        for col_idx, key in monthly_col_map.items():
            val = row[col_idx] if len(row) > col_idx else None
            if val is not None:
                try:
                    monthly[key] = int(val)
                except (ValueError, TypeError):
                    pass

        # Compute total from parsed monthly values if Excel formula was used
        total_raw = row[17] if len(row) > 17 else None
        total = _safe_int(total_raw) if not (
            total_raw and str(total_raw).startswith("=")
        ) else sum(monthly.values())

        records.append({
            "sr_no":                     len(records) + 1,
            "concern_id":                _safe_str(row[1] if len(row) > 1 else None),
            "concern":                   _safe_str(row[2] if len(row) > 2 else None),
            "type":                      _safe_str(row[3] if len(row) > 3 else None),
            "root_cause":                _safe_str(row[4] if len(row) > 4 else None),
            "action_plan":               _safe_str(row[5] if len(row) > 5 else None),
            "target_date":               _safe_date(row[6] if len(row) > 6 else None),
            "closure_date":              _safe_date(row[7] if len(row) > 7 else None),
            "ryg":                       _safe_str(row[8] if len(row) > 8 else None),
            "attri":                     _normalise_attri(_safe_str(row[9] if len(row) > 9 else None)),
            "comm":                      _safe_str(row[10] if len(row) > 10 else None),
            "line":                      _safe_str(row[11] if len(row) > 11 else None),
            "stage_no":                  _safe_str(row[12] if len(row) > 12 else None),
            "z_e":                       _safe_str(row[13] if len(row) > 13 else None),
            "attribution":               _safe_str(row[14] if len(row) > 14 else None),
            "part":                      _safe_str(row[15] if len(row) > 15 else None),
            "phenomena":                 _safe_str(row[16] if len(row) > 16 else None),
            "total_incidences":          total,
            "monthly_data":              json.dumps(monthly) if monthly else None,
            "field_defect_after_cutoff": _safe_int(row[field_defect_col] if len(row) > field_defect_col else None),
            "status_3m":                 _safe_str(row[status_3m_col] if len(row) > status_3m_col else None),
        })

    return records


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/upload", response_model=schemas.UploadResponse, status_code=201)
async def upload_excel(
    file: UploadFile = File(...),
    user_id: Optional[int] = Form(None),
    layout_id: Optional[int] = Form(None),
    connector: StateDBConnector = Depends(get_connector),
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted. Please open your file in Excel and save it as 'Excel Workbook (.xlsx)' before uploading.")

    file_bytes = await file.read()
    try:
        records = _parse_excel(file_bytes)
    except Exception as exc:
        logger.error(f"Excel parse error: {exc}")
        raise HTTPException(status_code=422, detail=f"Failed to parse Excel: {exc}")

    if not records:
        raise HTTPException(status_code=422, detail="No data rows found in the uploaded file")

    # Validate each row — skip invalid ones and collect skipped info
    valid_records = []
    skipped = []
    for i, rec in enumerate(records, start=2):  # row 1 is header, data starts at row 2
        reason = _validate_row(rec)
        if reason:
            skipped.append({"row_number": i, "reason": reason})
        else:
            valid_records.append(rec)

    if not valid_records:
        raise HTTPException(
            status_code=422,
            detail=f"All rows were invalid and skipped. No records imported.",
        )

    # Full replace for this user+layout scope, then insert fresh ones
    connector.execute_update(
        InputRecordQueries.DELETE_ALL,
        {"user_id": user_id, "layout_id": layout_id},
    )

    for rec in valid_records:
        rec["user_id"] = user_id
        rec["layout_id"] = layout_id
        connector.execute_query(InputRecordQueries.CREATE, rec)

    skipped_out = [{"row_number": s["row_number"], "reason": s["reason"]} for s in skipped]
    return {
        "message": "Upload successful",
        "rows_imported": len(valid_records),
        "skipped_rows": skipped_out if skipped_out else None,
    }


@router.get("/records", response_model=List[schemas.InputRecordOut])
def list_records(
    user_id: Optional[int] = Query(None),
    layout_id: Optional[int] = Query(None),
    connector: StateDBConnector = Depends(get_connector),
):
    rows = connector.execute_query(
        InputRecordQueries.LIST_ALL,
        {"user_id": user_id, "layout_id": layout_id},
    )
    return [_row_to_dict(r) for r in rows]


@router.post("/records", response_model=schemas.InputRecordOut, status_code=201)
def create_record(
    payload: schemas.InputRecordCreate,
    user_id: Optional[int] = Query(None),
    layout_id: Optional[int] = Query(None),
    connector: StateDBConnector = Depends(get_connector),
):
    data = payload.model_dump()
    # Validate constrained fields
    reason = _validate_row(data)
    if reason:
        raise HTTPException(status_code=422, detail=f"Validation error: {reason}")

    data["user_id"] = user_id
    data["layout_id"] = layout_id

    # Auto-assign SR No as max existing + 1
    max_rows = connector.execute_query(
        InputRecordQueries.GET_MAX_SR_NO,
        {"user_id": user_id, "layout_id": layout_id},
    )
    max_sr = max_rows[0][0] if max_rows else 0
    data["sr_no"] = (max_sr or 0) + 1

    # Ensure all required keys present (field_defect_after_cutoff always null from form)
    for key in ["concern_id", "concern", "type", "root_cause", "action_plan",
                "target_date", "closure_date", "ryg", "attri", "comm", "line", "stage_no",
                "z_e", "attribution", "part", "phenomena", "total_incidences",
                "monthly_data", "field_defect_after_cutoff", "status_3m"]:
        data.setdefault(key, None)

    rows = connector.execute_query(InputRecordQueries.CREATE, data)
    if not rows:
        raise HTTPException(status_code=500, detail="Failed to create record")
    return _row_to_dict(rows[0])


_IR_RETURNING = (
    "id, user_id, layout_id, sr_no, concern_id, concern, type, root_cause, action_plan, "
    "target_date, closure_date, ryg, attri, comm, line, stage_no, z_e, attribution, "
    "part, phenomena, total_incidences, monthly_data, field_defect_after_cutoff, "
    "status_3m, created_at, updated_at"
)


@router.put("/records/{record_id}", response_model=schemas.InputRecordOut)
def update_record(
    record_id: int,
    payload: schemas.InputRecordUpdate,
    connector: StateDBConnector = Depends(get_connector),
):
    exists = connector.execute_query(
        InputRecordQueries.CHECK_EXISTS, {"record_id": record_id}
    )
    if not exists:
        raise HTTPException(status_code=404, detail="Record not found")

    # Use only explicitly sent fields so that null genuinely clears the column
    # (COALESCE in the static query would silently keep the old value).
    data = payload.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(status_code=400, detail="No fields provided to update")

    set_clause = ", ".join(f"{col} = :{col}" for col in data)
    query = (
        f"UPDATE input_records SET {set_clause}, updated_at = NOW() "
        f"WHERE id = :record_id RETURNING {_IR_RETURNING}"
    )
    data["record_id"] = record_id
    rows = connector.execute_query(query, data)
    if not rows:
        raise HTTPException(status_code=500, detail="Failed to update record")
    return _row_to_dict(rows[0])


# ── Download endpoint ─────────────────────────────────────────────────────────

HEADER_ROW = [
    "Sr.No", "Concern ID", "Concern", "Type", "Root Cause", "Action Plan",
    "Target Date", "Closure Date", "RYG", "Attri.", "Comm", "Line", "Stage No",
    "Z/E", "ATTRIBUTION", "Part", "Phenomena", "Total Incidences",
    "Jan-24", "Feb-24", "Mar-24", "Apr-24", "May-24", "Jun-24",
    "Jul-24", "Aug-24", "Sep-24", "Oct-24", "Nov-24", "Dec-24",
    "Jan-25", "Feb-25", "Mar-25", "Apr-25", "May-25", "Jun-25",
    "Jul-25", "Aug-25", "Sep-25", "Oct-25", "Nov-25", "Dec-25",
    "Jan-26", "Feb-26", "Mar-26",
    "Field defect after cut off", "Status\n(3 Month basis)",
]


@router.get("/download")
def download_excel(
    user_id: Optional[int] = Query(None),
    layout_id: Optional[int] = Query(None),
    connector: StateDBConnector = Depends(get_connector),
):
    rows = connector.execute_query(
        InputRecordQueries.LIST_ALL,
        {"user_id": user_id, "layout_id": layout_id},
    )
    records = [_row_to_dict(r) for r in rows]

    # Collect all month keys present in the data, sorted chronologically
    all_monthly_keys: list[str] = []
    all_monthly_set: set[str] = set()
    for rec in records:
        if rec.get("monthly_data"):
            try:
                m = json.loads(rec["monthly_data"])
                all_monthly_set.update(m.keys())
            except (ValueError, TypeError):
                pass
    # Merge hardcoded base keys + any extra keys from data, keep sorted
    all_monthly_keys = sorted(all_monthly_set | set(MONTHLY_KEYS))

    # Build dynamic header row
    month_labels = []
    for key in all_monthly_keys:
        year, month = key.split("-")
        month_labels.append(
            f"{['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec'][int(month)-1]}-{year[2:]}"
        )

    dynamic_header = [
        "Sr.No", "Concern ID", "Concern", "Type", "Root Cause", "Action Plan",
        "Target Date", "Closure Date", "RYG", "Attri.", "Comm", "Line", "Stage No",
        "Z/E", "ATTRIBUTION", "Part", "Phenomena", "Total Incidences",
        *month_labels,
        "Field defect after cut off", "Status\n(3 Month basis)",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Data"

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(dynamic_header)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 30

    # Data rows
    for rec in records:
        monthly = {}
        if rec.get("monthly_data"):
            try:
                monthly = json.loads(rec["monthly_data"])
            except (ValueError, TypeError):
                pass

        row = [
            rec.get("sr_no"),
            rec.get("concern_id"),
            rec.get("concern"),
            rec.get("type"),
            rec.get("root_cause"),
            rec.get("action_plan"),
            rec.get("target_date"),
            rec.get("closure_date"),
            rec.get("ryg"),
            rec.get("attri"),
            rec.get("comm"),
            rec.get("line"),
            rec.get("stage_no"),
            rec.get("z_e"),
            rec.get("attribution"),
            rec.get("part"),
            rec.get("phenomena"),
            rec.get("total_incidences"),
        ]
        for key in all_monthly_keys:
            row.append(monthly.get(key))
        row.append(rec.get("field_defect_after_cutoff"))
        row.append(rec.get("status_3m"))
        ws.append(row)

    # Auto-fit column widths
    col_widths = [8, 16, 40, 8, 35, 35, 14, 14, 7, 20, 16, 18, 12, 7, 14, 22, 22, 10]
    col_widths += [8] * len(all_monthly_keys)
    col_widths += [22, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "z_stage_master_data.xlsx"
    if layout_id is not None:
        filename = f"z_stage_master_data_layout_{layout_id}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
