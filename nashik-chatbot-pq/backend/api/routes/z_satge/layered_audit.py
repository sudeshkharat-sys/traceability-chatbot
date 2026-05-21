import io
import logging
import re
import datetime
from typing import List, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

from app.connectors.state_db_connector import StateDBConnector
from app.connectors.database import get_connector
from app.queries import LayeredAuditQueries, LayeredAuditAdherenceQueries
import backend.models.schemas.z_stage_schemas as schemas

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/layered-audit", tags=["layered_audit"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _row_to_dict(row) -> dict:
    return dict(row._mapping)


def _safe_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    # strip non-breaking spaces and zero-width chars
    s = re.sub(r'[\u00a0\u200b\u200c\u200d\u200e\u200f\ufeff\u00ad]', '', s)
    if s.startswith("=") or not s:
        return None
    return s


def _safe_date(value) -> str | None:
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s or s.startswith("=") or s == "00:00:00":
        return None
    return s


_DATE_FORMATS = (
    "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y",
    "%d.%m.%Y", "%Y/%m/%d", "%d-%b-%Y", "%d %b %Y",
)

def _strict_date(value) -> str | None:
    """Strict date parser for audit_date: only accepts real dates.
    Rejects 'NA', 'N/A', free-text strings — stores None (blank) instead."""
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s or s.startswith("=") or s == "00:00:00":
        return None
    # Try all known date formats; only store if a valid date can be parsed
    token = s[:10]  # ignore any time component
    for fmt in _DATE_FORMATS:
        try:
            dt = datetime.datetime.strptime(token, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Could not parse → leave blank (NA, N/A, invalid text, etc.)
    return None


# ── Parse Layered Audit Excel ─────────────────────────────────────────────────
# Expected columns (1-indexed):
#  1=Model, 2=Sr.No, 3=Date, 4=Station ID, 5=Workstation, 6=Auditor,
#  7=NC's, 8=Action Plan, 9=4M, 10=Responsibility, 11=target Date, 12=Status

_LA_COL_MAP = {
    "model":          "model",
    "sr.no":          "sr_no",
    "sr no":          "sr_no",
    "date":           "date_col",
    "station id":     "station_id",
    "workstation":    "workstation",
    "auditor":        "auditor",
    "nc's":           "ncs",
    "ncs":            "ncs",
    "action plan":    "action_plan",
    "4m":             "four_m",
    "responsibility": "responsibility",
    "target date":    "target_date",
    "status":         "status",
}


def _parse_layered_audit(file_bytes: bytes) -> list[dict]:
    """Original positional parser — used for Replace All uploads (full template expected)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # skip header
    records = []
    for row in rows_iter:
        if not any(v is not None for v in row):
            continue
        records.append({
            "model":          _safe_str(row[0] if len(row) > 0 else None),
            "sr_no":          _safe_str(row[1] if len(row) > 1 else None),
            "date_col":       _safe_date(row[2] if len(row) > 2 else None),
            "station_id":     _safe_str(row[3] if len(row) > 3 else None),
            "workstation":    _safe_str(row[4] if len(row) > 4 else None),
            "auditor":        _safe_str(row[5] if len(row) > 5 else None),
            "ncs":            _safe_str(row[6] if len(row) > 6 else None),
            "action_plan":    _safe_str(row[7] if len(row) > 7 else None),
            "four_m":         _safe_str(row[8] if len(row) > 8 else None),
            "responsibility": _safe_str(row[9] if len(row) > 9 else None),
            "target_date":    _safe_date(row[10] if len(row) > 10 else None),
            "status":         _safe_str(row[11] if len(row) > 11 else None),
        })
    return records


def _parse_layered_audit_append(file_bytes: bytes) -> list[dict]:
    """Header-name-based parser — used for Append mode (columns may be missing/reordered)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []

    field_col: dict[str, int] = {}
    for col_idx, cell_val in enumerate(header):
        if isinstance(cell_val, str):
            clean = cell_val.strip().lower().replace("\n", " ").replace("  ", " ")
            field = _LA_COL_MAP.get(clean)
            if field and field not in field_col:
                field_col[field] = col_idx

    def _col(row, field):
        idx = field_col.get(field)
        return row[idx] if idx is not None and len(row) > idx else None

    records = []
    for row in rows_iter:
        if not any(v is not None for v in row):
            continue
        records.append({
            "model":          _safe_str(_col(row, "model")),
            "sr_no":          _safe_str(_col(row, "sr_no")),
            "date_col":       _safe_date(_col(row, "date_col")),
            "station_id":     _safe_str(_col(row, "station_id")),
            "workstation":    _safe_str(_col(row, "workstation")),
            "auditor":        _safe_str(_col(row, "auditor")),
            "ncs":            _safe_str(_col(row, "ncs")),
            "action_plan":    _safe_str(_col(row, "action_plan")),
            "four_m":         _safe_str(_col(row, "four_m")),
            "responsibility": _safe_str(_col(row, "responsibility")),
            "target_date":    _safe_date(_col(row, "target_date")),
            "status":         _safe_str(_col(row, "status")),
        })
    return records


# ── Parse Layered Audit Adherence Excel ──────────────────────────────────────
# Expected columns (1-indexed):
#  1=Stage No, 2=Stage Name, 3=Auditor, 4=Audit Date

_LAA_COL_MAP = {
    "stage no":   "stage_no",
    "stage name": "stage_name",
    "auditor":    "auditor",
    "audit date": "audit_date",
}


def _parse_layered_audit_adherence(file_bytes: bytes) -> list[dict]:
    """Original positional parser — used for Replace All uploads (full template expected)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # skip header
    records = []
    for row in rows_iter:
        if not any(v is not None for v in row):
            continue
        records.append({
            "stage_no":   _safe_str(row[0] if len(row) > 0 else None),
            "stage_name": _safe_str(row[1] if len(row) > 1 else None),
            "auditor":    _safe_str(row[2] if len(row) > 2 else None),
            "audit_date": _strict_date(row[3] if len(row) > 3 else None),
        })
    return records


def _parse_layered_audit_adherence_append(file_bytes: bytes) -> list[dict]:
    """Header-name-based parser — used for Append mode (columns may be missing/reordered)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []

    field_col: dict[str, int] = {}
    for col_idx, cell_val in enumerate(header):
        if isinstance(cell_val, str):
            clean = cell_val.strip().lower().replace("\n", " ").replace("  ", " ")
            field = _LAA_COL_MAP.get(clean)
            if field and field not in field_col:
                field_col[field] = col_idx

    def _col(row, field):
        idx = field_col.get(field)
        return row[idx] if idx is not None and len(row) > idx else None

    records = []
    for row in rows_iter:
        if not any(v is not None for v in row):
            continue
        records.append({
            "stage_no":   _safe_str(_col(row, "stage_no")),
            "stage_name": _safe_str(_col(row, "stage_name")),
            "auditor":    _safe_str(_col(row, "auditor")),
            "audit_date": _strict_date(_col(row, "audit_date")),
        })
    return records


# ── Layered Audit endpoints ───────────────────────────────────────────────────

@router.post("/records", response_model=schemas.LayeredAuditOut, status_code=201)
def create_layered_audit(
    payload: schemas.LayeredAuditCreate,
    user_id: Optional[int] = Query(None),
    layout_id: Optional[int] = Query(None),
    connector: StateDBConnector = Depends(get_connector),
):
    data = payload.model_dump()
    data["user_id"] = user_id
    data["layout_id"] = layout_id
    for key in ["model", "sr_no", "date_col", "station_id", "workstation", "auditor",
                "ncs", "action_plan", "four_m", "responsibility", "target_date", "status"]:
        data.setdefault(key, None)
    rows = connector.execute_query(LayeredAuditQueries.CREATE, data)
    if not rows:
        raise HTTPException(status_code=500, detail="Failed to create record")
    return _row_to_dict(rows[0])


@router.post("/upload", response_model=schemas.UploadResponse, status_code=201)
async def upload_layered_audit(
    file: UploadFile = File(...),
    user_id: Optional[int] = Form(None),
    layout_id: Optional[int] = Form(None),
    mode: str = Form("replace"),
    connector: StateDBConnector = Depends(get_connector),
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted. Please open your file in Excel and save it as 'Excel Workbook (.xlsx)' before uploading.")

    file_bytes = await file.read()
    try:
        records = _parse_layered_audit_append(file_bytes) if mode == "append" else _parse_layered_audit(file_bytes)
    except Exception as exc:
        logger.error(f"Layered Audit parse error: {exc}")
        raise HTTPException(status_code=422, detail=f"Failed to parse Excel: {exc}")

    if not records:
        raise HTTPException(status_code=422, detail="No data rows found in the uploaded file")

    if mode != "append":
        connector.execute_update(
            LayeredAuditQueries.DELETE_ALL,
            {"user_id": user_id, "layout_id": layout_id},
        )

    for rec in records:
        rec["user_id"] = user_id
        rec["layout_id"] = layout_id
        connector.execute_query(LayeredAuditQueries.CREATE, rec)

    action_word = "appended" if mode == "append" else "imported"
    return {"message": f"Layered Audit upload successful — {len(records)} rows {action_word}", "rows_imported": len(records)}


@router.get("/records", response_model=List[schemas.LayeredAuditOut])
def list_layered_audit(
    user_id: Optional[int] = Query(None),
    layout_id: Optional[int] = Query(None),
    connector: StateDBConnector = Depends(get_connector),
):
    rows = connector.execute_query(
        LayeredAuditQueries.LIST_ALL,
        {"user_id": user_id, "layout_id": layout_id},
    )
    return [_row_to_dict(r) for r in rows]


@router.delete("/records/{record_id}", status_code=204)
def delete_layered_audit(
    record_id: int,
    connector: StateDBConnector = Depends(get_connector),
):
    exists = connector.execute_query(
        LayeredAuditQueries.CHECK_EXISTS, {"record_id": record_id}
    )
    if not exists:
        raise HTTPException(status_code=404, detail="Record not found")
    connector.execute_update(LayeredAuditQueries.DELETE_BY_ID, {"record_id": record_id})


_LA_RETURNING = (
    "id, user_id, layout_id, model, sr_no, date_col, station_id, "
    "workstation, auditor, ncs, action_plan, four_m, responsibility, "
    "target_date, status, created_at, updated_at"
)


@router.put("/records/{record_id}", response_model=schemas.LayeredAuditOut)
def update_layered_audit(
    record_id: int,
    payload: schemas.LayeredAuditUpdate,
    connector: StateDBConnector = Depends(get_connector),
):
    exists = connector.execute_query(
        LayeredAuditQueries.CHECK_EXISTS, {"record_id": record_id}
    )
    if not exists:
        raise HTTPException(status_code=404, detail="Record not found")

    # Use only the fields that were explicitly sent (exclude_unset=True) so that
    # sending null actually clears the column instead of being swallowed by COALESCE.
    data = payload.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(status_code=400, detail="No fields provided to update")

    set_clause = ", ".join(f"{col} = :{col}" for col in data)
    query = (
        f"UPDATE layered_audit SET {set_clause}, updated_at = NOW() "
        f"WHERE id = :record_id RETURNING {_LA_RETURNING}"
    )
    data["record_id"] = record_id
    rows = connector.execute_query(query, data)
    if not rows:
        raise HTTPException(status_code=500, detail="Failed to update record")
    return _row_to_dict(rows[0])


# ── Layered Audit Adherence endpoints ────────────────────────────────────────

@router.post("/adherence/records", response_model=schemas.LayeredAuditAdherenceOut, status_code=201)
def create_layered_audit_adherence(
    payload: schemas.LayeredAuditAdherenceCreate,
    user_id: Optional[int] = Query(None),
    layout_id: Optional[int] = Query(None),
    connector: StateDBConnector = Depends(get_connector),
):
    data = payload.model_dump()
    # Enforce date validation for audit_date
    if data.get("audit_date") is not None:
        data["audit_date"] = _strict_date(data["audit_date"])
    data["user_id"] = user_id
    data["layout_id"] = layout_id
    for key in ["stage_no", "stage_name", "auditor", "audit_date"]:
        data.setdefault(key, None)
    rows = connector.execute_query(LayeredAuditAdherenceQueries.CREATE, data)
    if not rows:
        raise HTTPException(status_code=500, detail="Failed to create record")
    return _row_to_dict(rows[0])


@router.post("/adherence/upload", response_model=schemas.UploadResponse, status_code=201)
async def upload_layered_audit_adherence(
    file: UploadFile = File(...),
    user_id: Optional[int] = Form(None),
    layout_id: Optional[int] = Form(None),
    mode: str = Form("replace"),
    connector: StateDBConnector = Depends(get_connector),
):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted. Please open your file in Excel and save it as 'Excel Workbook (.xlsx)' before uploading.")

    file_bytes = await file.read()
    try:
        records = _parse_layered_audit_adherence_append(file_bytes) if mode == "append" else _parse_layered_audit_adherence(file_bytes)
    except Exception as exc:
        logger.error(f"Layered Audit Adherence parse error: {exc}")
        raise HTTPException(status_code=422, detail=f"Failed to parse Excel: {exc}")

    if not records:
        raise HTTPException(status_code=422, detail="No data rows found in the uploaded file")

    if mode != "append":
        connector.execute_update(
            LayeredAuditAdherenceQueries.DELETE_ALL,
            {"user_id": user_id, "layout_id": layout_id},
        )

    for rec in records:
        rec["user_id"] = user_id
        rec["layout_id"] = layout_id
        connector.execute_query(LayeredAuditAdherenceQueries.CREATE, rec)

    action_word = "appended" if mode == "append" else "imported"
    return {"message": f"Layered Audit Adherence upload successful — {len(records)} rows {action_word}", "rows_imported": len(records)}


@router.get("/adherence/records", response_model=List[schemas.LayeredAuditAdherenceOut])
def list_layered_audit_adherence(
    user_id: Optional[int] = Query(None),
    layout_id: Optional[int] = Query(None),
    connector: StateDBConnector = Depends(get_connector),
):
    rows = connector.execute_query(
        LayeredAuditAdherenceQueries.LIST_ALL,
        {"user_id": user_id, "layout_id": layout_id},
    )
    return [_row_to_dict(r) for r in rows]


@router.delete("/adherence/records/{record_id}", status_code=204)
def delete_layered_audit_adherence(
    record_id: int,
    connector: StateDBConnector = Depends(get_connector),
):
    exists = connector.execute_query(
        LayeredAuditAdherenceQueries.CHECK_EXISTS, {"record_id": record_id}
    )
    if not exists:
        raise HTTPException(status_code=404, detail="Record not found")
    connector.execute_update(LayeredAuditAdherenceQueries.DELETE_BY_ID, {"record_id": record_id})


_LAA_RETURNING = (
    "id, user_id, layout_id, stage_no, stage_name, auditor, audit_date, "
    "created_at, updated_at"
)


@router.put("/adherence/records/{record_id}", response_model=schemas.LayeredAuditAdherenceOut)
def update_layered_audit_adherence(
    record_id: int,
    payload: schemas.LayeredAuditAdherenceUpdate,
    connector: StateDBConnector = Depends(get_connector),
):
    exists = connector.execute_query(
        LayeredAuditAdherenceQueries.CHECK_EXISTS, {"record_id": record_id}
    )
    if not exists:
        raise HTTPException(status_code=404, detail="Record not found")

    data = payload.model_dump(exclude_unset=True)
    if not data:
        raise HTTPException(status_code=400, detail="No fields provided to update")

    # Enforce date validation for audit_date when it is explicitly provided
    if "audit_date" in data and data["audit_date"] is not None:
        data["audit_date"] = _strict_date(data["audit_date"])

    set_clause = ", ".join(f"{col} = :{col}" for col in data)
    query = (
        f"UPDATE layered_audit_adherence SET {set_clause}, updated_at = NOW() "
        f"WHERE id = :record_id RETURNING {_LAA_RETURNING}"
    )
    data["record_id"] = record_id
    rows = connector.execute_query(query, data)
    if not rows:
        raise HTTPException(status_code=500, detail="Failed to update record")
    return _row_to_dict(rows[0])


# ── Download endpoints ────────────────────────────────────────────────────────

_LA_HEADER = ["Model", "Sr.No", "Date", "Station ID", "Workstation", "Auditor",
               "NC's", "Action Plan", "4M", "Responsibility", "Target Date", "Status"]
_LA_FIELDS = ["model", "sr_no", "date_col", "station_id", "workstation", "auditor",
               "ncs", "action_plan", "four_m", "responsibility", "target_date", "status"]

_LAA_HEADER = ["Stage No", "Stage Name", "Auditor", "Audit Date"]
_LAA_FIELDS = ["stage_no", "stage_name", "auditor", "audit_date"]


def _make_excel(header: list, fields: list, records: list, sheet_title: str) -> io.BytesIO:
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.append(header)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 28
    for rec in records:
        ws.append([rec.get(f) for f in fields])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/download")
def download_layered_audit(
    user_id: Optional[int] = Query(None),
    layout_id: Optional[int] = Query(None),
    connector: StateDBConnector = Depends(get_connector),
):
    rows = connector.execute_query(
        LayeredAuditQueries.LIST_ALL,
        {"user_id": user_id, "layout_id": layout_id},
    )
    records = [_row_to_dict(r) for r in rows]
    buf = _make_excel(_LA_HEADER, _LA_FIELDS, records, "Layered Audit")
    filename = f"layered_audit_layout_{layout_id}.xlsx" if layout_id else "layered_audit.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/adherence/download")
def download_layered_audit_adherence(
    user_id: Optional[int] = Query(None),
    layout_id: Optional[int] = Query(None),
    connector: StateDBConnector = Depends(get_connector),
):
    rows = connector.execute_query(
        LayeredAuditAdherenceQueries.LIST_ALL,
        {"user_id": user_id, "layout_id": layout_id},
    )
    records = [_row_to_dict(r) for r in rows]
    buf = _make_excel(_LAA_HEADER, _LAA_FIELDS, records, "Audit Adherence")
    filename = f"audit_adherence_layout_{layout_id}.xlsx" if layout_id else "audit_adherence.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
